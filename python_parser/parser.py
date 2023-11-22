import openpyxl as op
import json

book = op.load_workbook('parsed_data.xlsx')
sheet_UF = book['UF_Comms']
sheet_UFR = book['UFR Comms']

moves_dict_uf = dict()
moves_dict_ufr = dict()

start_row, end_row = 2, 23
start_col, end_col = 2, 23

keys_uf = []
for key in sheet_UF['B1:W1'][0]:
    keys_uf.append(key.value)

keys_ufr = []

for key in sheet_UFR['B1:V1'][0]:
    keys_ufr.append(key.value)


offset = 2

for row in range(start_row, end_row+1):
    for col in range(start_col, end_col+1):
        UF_cell = sheet_UF.cell(row=row, column=col)
        UF_bg_color = UF_cell.fill.start_color.rgb if UF_cell.fill.start_color else 'FFFFFFFF'
        cell_key = keys_uf[UF_cell.row -offset][0]+keys_uf[UF_cell.column-offset][0]
        moves_dict_uf[cell_key] = {
            'UF_value':UF_cell.value,
            'UF_color':f'#{UF_bg_color[2::]}',
        }

with open('data_uf.json', 'w') as json_file:
    json.dump(moves_dict_uf, json_file)


start_row, end_row = 2, 22
start_col, end_col = 2, 22

for row in range(start_row, end_row+1):
    for col in range(start_col, end_col+1):
        UFR_cell = sheet_UFR.cell(row=row, column=col)
        UFR_bg_color = UFR_cell.fill.start_color.rgb if UFR_cell.fill.start_color else 'FFFFFFFF'
        cell_key = keys_ufr[UFR_cell.row -offset][0]+keys_ufr[UFR_cell.column-offset][0]
        moves_dict_ufr[cell_key] = {
            'UFR_value': UFR_cell.value,
            'UFR_color': f'#{UFR_bg_color[2::]}'
        }

with open('data_ufr.json', 'w') as json_file:
    json.dump(moves_dict_ufr, json_file)



book.close()